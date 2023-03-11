import pandas as pd
import numpy as np
import openpyxl as op
import string as s
import os

if __name__ == '__main__':
    run_class = True
    list_of_classes = []
    marks_path = "Individual Marks"
    while run_class:

        class_code = input("Please enter the name or code of your class.\n")

        run_class_work = True
        class_list_of_work = []
        class_list_of_weight = []
        list_mark_input = []

        while run_class_work == True:
            class_work = input("Enter an assesment or classwork here.\n")
            class_weight = input("Enter the percentage weight of that task (Numbers from 1 - 100 are acceptable here). Just make sure your final values sum to 100.\n")
            class_list_of_work.append(class_work)
            class_list_of_weight.append(class_weight)
            list_mark_input.append(np.nan)
            ask_to_continue_class_work = input("Would you like to add more assesments/classwork? Y for yes, N for no.\n")
            if ask_to_continue_class_work == 'Y': continue
            else: run_class_work = False
        
        class_data = {class_code: class_list_of_work, "Weight": class_list_of_weight, "Enter Mark Here": list_mark_input, 
        "Total Points": list_mark_input, "Current Mark": list_mark_input}
        class_df = pd.DataFrame(data=class_data)

        weight_letter = s.ascii_uppercase[class_df.columns.get_loc("Weight")]
        mark_letter = s.ascii_uppercase[class_df.columns.get_loc("Enter Mark Here")]
        total_points_letter = s.ascii_uppercase[class_df.columns.get_loc("Total Points")]
        mark_entry_letter = s.ascii_uppercase[class_df.columns.get_loc("Enter Mark Here")]
        j = 2; k = j

        for i in range(2, len(class_list_of_work) + 2):
            class_df.loc[i - 2,"Total Points"] = f"=PRODUCT({weight_letter + str(i)},{mark_letter + str(i)})"; k+=1

        class_df.loc[0, "Current Mark"] = f"=SUMIF({mark_entry_letter + str(j)}:{mark_entry_letter + str(k)},\">0\",{total_points_letter + str(j)}:{total_points_letter + str(k)})/SUMIF({mark_entry_letter + str(j)}:{mark_entry_letter + str(k)},\">0\",{weight_letter + str(j)}:{weight_letter + str(k)})"
        class_df.to_excel(f"{marks_path}\\{class_code}_Marks.xlsx")

        wb_class = op.load_workbook(f"{marks_path}\\{class_code}_Marks.xlsx")
        ws_class = wb_class.active
        ws_class.delete_cols(1)
        ws_class.title = class_code

        wb_class.save(filename = f"{marks_path}\\{class_code}_Marks.xlsx")

        list_of_classes.append(class_df)
        add_another_class = input("Would you like to add another class? Y for yes, N for no.\n")
        if add_another_class ==  "Y": continue
        else: run_class = False

    count = 0
    marks_wb = op.Workbook()
    for excel_file in os.listdir(marks_path):
        excel_file_wb = op.load_workbook(f"{marks_path}\\{excel_file}")
        excel_file_ws = excel_file_wb.worksheets[0]
        course_title = excel_file.split("_")[0]

        marks_wb.create_sheet(title=course_title,index=count)
        marks_ws = marks_wb.worksheets[count]

        row_max = excel_file_ws.max_row
        column_max = excel_file_ws.max_column

        for i in range(1, row_max + 1):
            for j in range(1, column_max + 1):
                marks_ws.cell(row=i, column=j).value = excel_file_ws.cell(row=i, column=j).value
        count += 1
    to_delete = marks_wb.get_sheet_by_name("Sheet")
    marks_wb.remove_sheet(to_delete)
    marks_wb.save(filename="Marks.xlsx")

    print('Your marks spreadsheet has been generated.')